"""Importa o aggiorna Sponsor/clienti da un file Excel."""
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


COLONNE_RICHIESTE = [
    "ragione_sociale", "nome_commerciale", "partita_iva", "codice_fiscale",
    "codice_sdi", "pec", "indirizzo", "citta", "cap", "provincia", "paese",
    "settore", "sito_web", "note",
    "referente_nome", "referente_email", "referente_telefono", "referente_ruolo",
]
COLONNE_OBBLIGATORIE = ["ragione_sociale"]


class Command(BaseCommand):
    help = "Importa o aggiorna Sponsor/clienti da un file Excel (vedi template_sponsor.xlsx)."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Percorso del file Excel.")
        parser.add_argument("--dry-run", action="store_true",
                            help="Simula senza scrivere nulla nel DB.")

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError("openpyxl non installato. Installa con: pip install openpyxl")
        from sponsors.models import Sponsor, Contact

        path = Path(opts["file"]).expanduser()
        if not path.exists():
            raise CommandError(f"File non trovato: {path}")

        dry = opts["dry_run"]
        self.stdout.write(f"Lettura {path.name}" + (" [DRY-RUN]" if dry else ""))

        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f"Impossibile aprire il file Excel: {e}")

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("Il foglio e' vuoto.")

        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        mancanti = [c for c in COLONNE_OBBLIGATORIE if c not in header]
        if mancanti:
            raise CommandError(
                f"Colonne obbligatorie mancanti nell'intestazione: {mancanti}. "
                f"Atteso almeno: {COLONNE_OBBLIGATORIE}")
        col_idx = {h: i for i, h in enumerate(header)}

        n_create = n_update = n_err = n_contatti = 0
        errori = []

        for n_riga, row in enumerate(rows[1:], start=2):
            if all(v is None or v == "" for v in row):
                continue

            def G(nome):
                i = col_idx.get(nome)
                v = row[i] if i is not None and i < len(row) else None
                return ("" if v is None else str(v)).strip()

            try:
                rs = G("ragione_sociale")
                if not rs:
                    raise ValueError("ragione_sociale obbligatoria")
                piva = G("partita_iva")

                sponsor = None
                if piva:
                    sponsor = Sponsor.objects.filter(vat_number=piva).first()
                if sponsor is None:
                    sponsor = Sponsor.objects.filter(legal_name__iexact=rs).first()
                creato = sponsor is None

                if dry:
                    self.stdout.write(
                        f"  {n_riga:>4}: {'CREEREBBE' if creato else 'AGGIORNEREBBE'} '{rs}'"
                        + (f" (P.IVA {piva})" if piva else ""))
                    if creato:
                        n_create += 1
                    else:
                        n_update += 1
                    continue

                with transaction.atomic():
                    if sponsor is None:
                        sponsor = Sponsor(legal_name=rs)
                    else:
                        sponsor.legal_name = rs

                    mappa = {
                        "display_name": "nome_commerciale",
                        "vat_number": "partita_iva",
                        "tax_code": "codice_fiscale",
                        "sdi_code": "codice_sdi",
                        "pec_email": "pec",
                        "address_street": "indirizzo",
                        "address_city": "citta",
                        "address_zip": "cap",
                        "address_province": "provincia",
                        "industry": "settore",
                        "website": "sito_web",
                        "notes": "note",
                    }
                    for attr, col in mappa.items():
                        val = G(col)
                        if val:
                            setattr(sponsor, attr, val)
                    paese = G("paese")
                    if paese:
                        sponsor.address_country = paese.upper()[:2]
                    sponsor.save()

                    if creato:
                        n_create += 1
                        self.stdout.write(f"  {n_riga:>4}: + CREATO '{rs}'")
                    else:
                        n_update += 1
                        self.stdout.write(f"  {n_riga:>4}: ~ AGGIORNATO '{rs}'")

                    ref_email = G("referente_email")
                    ref_nome = G("referente_nome")
                    if ref_email and ref_nome:
                        contact = Contact.objects.filter(
                            sponsor=sponsor, email__iexact=ref_email).first()
                        if contact is None:
                            contact = Contact(sponsor=sponsor, email=ref_email, full_name=ref_nome)
                        else:
                            contact.full_name = ref_nome
                        tel = G("referente_telefono")
                        ruolo = G("referente_ruolo")
                        if tel:
                            contact.phone = tel
                        if ruolo:
                            contact.job_title = ruolo
                        if not sponsor.contacts.filter(is_primary=True).exists():
                            contact.is_primary = True
                        contact.save()
                        n_contatti += 1
                        self.stdout.write(f"       + referente '{ref_nome}' <{ref_email}>")
                    elif ref_email or ref_nome:
                        self.stdout.write(self.style.WARNING(
                            f"  {n_riga:>4}: referente ignorato (servono SIA nome SIA email)"))

            except Exception as e:
                n_err += 1
                msg = f"riga {n_riga}: ERRORE {e}"
                errori.append(msg)
                self.stdout.write(self.style.ERROR("  " + msg))

        self.stdout.write("")
        riepilogo = (f"Fatto: {n_create} sponsor creati, {n_update} aggiornati, "
                     f"{n_contatti} referenti collegati, {n_err} errori.")
        if dry:
            riepilogo = "[DRY-RUN] " + riepilogo + " (nessun salvataggio)"
        if n_err:
            self.stdout.write(self.style.WARNING(riepilogo))
        else:
            self.stdout.write(self.style.SUCCESS(riepilogo))
