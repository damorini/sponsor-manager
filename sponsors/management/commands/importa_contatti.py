"""Importa o aggiorna Contatti (referenti) da un file Excel."""
import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


COLONNE_RICHIESTE = [
    "sponsor_partita_iva", "sponsor_ragione_sociale",
    "cognome", "nome", "email", "telefono", "ruolo_aziendale",
    "ruoli_funzionali", "principale", "consenso_marketing", "lingua", "note",
]
# Solo 'email' a livello header: il nome (cognome o nome_completo) si valida riga per riga,
# cosi' funzionano sia i nuovi file (cognome/nome) sia i vecchi (nome_completo).
COLONNE_OBBLIGATORIE = ["email"]

ROLE_MAP = {
    "signer": "signer", "firmatario": "signer",
    "marketing": "marketing",
    "finance": "finance", "amministrazione": "finance", "amm": "finance",
    "operational": "operational", "operativo": "operational",
    "cc": "cc",
    "educational": "educational", "educational manager": "educational",
    "educational_manager": "educational",
}


def _norm_bool(v, default=False):
    if v is None or v == "":
        return default
    s = str(v).strip().lower()
    if s in {"s", "si", "sì", "yes", "y", "1", "true", "x", "✓"}:
        return True
    if s in {"n", "no", "0", "false", "-"}:
        return False
    return default


def _parse_roles(raw):
    if not raw:
        return [], []
    tokens = [t.strip().lower() for t in re.split(r"[,;/]+", str(raw)) if t.strip()]
    validi, ignorati = [], []
    for t in tokens:
        code = ROLE_MAP.get(t)
        if code and code not in validi:
            validi.append(code)
        elif not code:
            ignorati.append(t)
    return validi, ignorati


class Command(BaseCommand):
    help = "Importa o aggiorna Contatti da un file Excel (vedi template_contatti.xlsx)."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Percorso del file Excel.")
        parser.add_argument("--dry-run", action="store_true",
                            help="Simula senza scrivere nulla nel DB.")

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError("openpyxl non installato. Installa con: pip install openpyxl")
        from sponsors.models import Sponsor, Contact

        path = Path(opts["file"]).expanduser()
        if not path.exists():
            raise CommandError(f"File non trovato: {path}")
        dry = opts["dry_run"]
        self.stdout.write(f"Lettura {path.name}" + (" [DRY-RUN]" if dry else ""))
        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f"Impossibile aprire il file Excel: {e}")
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("Il foglio e' vuoto.")
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        mancanti = [c for c in COLONNE_OBBLIGATORIE if c not in header]
        if mancanti:
            raise CommandError(f"Colonne obbligatorie mancanti nell'intestazione: {mancanti}.")
        col_idx = {h: i for i, h in enumerate(header)}

        n_create = n_update = n_err = 0
        errori = []
        for n_riga, row in enumerate(rows[1:], start=2):
            if all(v is None or v == "" for v in row):
                continue

            def G(nome):
                i = col_idx.get(nome)
                v = row[i] if i is not None and i < len(row) else None
                return ("" if v is None else str(v)).strip()

            try:
                cognome = G("cognome")
                nome_p = G("nome")
                nome_completo = G("nome_completo")  # compat vecchi file
                email = G("email")
                nome = nome_completo or (nome_p + " " + cognome).strip()
                if not (cognome or nome_completo) or not email:
                    raise ValueError("cognome (o nome_completo) ed email sono obbligatori")

                piva = G("sponsor_partita_iva")
                rs = G("sponsor_ragione_sociale")
                sponsor = None
                if piva:
                    sponsor = Sponsor.objects.filter(vat_number=piva).first()
                if sponsor is None and rs:
                    sponsor = Sponsor.objects.filter(legal_name__iexact=rs).first()
                if sponsor is None:
                    raise ValueError(
                        "sponsor non trovato (indica sponsor_partita_iva o "
                        "sponsor_ragione_sociale di uno sponsor gia' esistente)")

                contact = Contact.objects.filter(sponsor=sponsor, email__iexact=email).first()
                creato = contact is None

                if dry:
                    self.stdout.write(
                        f"  {n_riga:>4}: {'CREEREBBE' if creato else 'AGGIORNEREBBE'} "
                        f"'{nome}' <{email}> su '{sponsor.legal_name}'")
                    if creato:
                        n_create += 1
                    else:
                        n_update += 1
                    continue

                with transaction.atomic():
                    if contact is None:
                        contact = Contact(sponsor=sponsor, email=email)
                    # Nome/Cognome separati (il full_name si ricompone al salvataggio);
                    # se il file ha solo nome_completo, lo usiamo (verra' diviso).
                    if cognome or nome_p:
                        contact.last_name = cognome
                        contact.first_name = nome_p
                    elif nome_completo:
                        contact.full_name = nome_completo
                        contact.first_name = ''
                        contact.last_name = ''
                    tel = G("telefono")
                    ruolo = G("ruolo_aziendale")
                    if tel:
                        contact.phone = tel
                    if ruolo:
                        contact.job_title = ruolo
                    ruoli_raw = G("ruoli_funzionali")
                    if ruoli_raw:
                        validi, ignorati = _parse_roles(ruoli_raw)
                        contact.roles = validi
                        if ignorati:
                            self.stdout.write(self.style.WARNING(
                                f"  {n_riga:>4}: ruoli ignorati (non validi): {ignorati}"))
                    lingua = G("lingua").lower()
                    if lingua in {"it", "italiano"}:
                        contact.preferred_language = "it"
                    elif lingua in {"en", "inglese", "english"}:
                        contact.preferred_language = "en"
                    if G("consenso_marketing"):
                        contact.marketing_consent = _norm_bool(G("consenso_marketing"))
                    note = G("note")
                    if note:
                        contact.notes = note

                    principale_set = G("principale")
                    if principale_set:
                        contact.is_primary = _norm_bool(principale_set)
                    contact.save()
                    if principale_set and contact.is_primary:
                        sponsor.contacts.filter(is_primary=True).exclude(
                            pk=contact.pk).update(is_primary=False)

                    if creato:
                        n_create += 1
                        self.stdout.write(
                            f"  {n_riga:>4}: + CREATO '{nome}' <{email}> su '{sponsor.legal_name}'")
                    else:
                        n_update += 1
                        self.stdout.write(f"  {n_riga:>4}: ~ AGGIORNATO '{nome}' <{email}>")

            except Exception as e:
                n_err += 1
                msg = f"riga {n_riga}: ERRORE {e}"
                errori.append(msg)
                self.stdout.write(self.style.ERROR("  " + msg))

        self.stdout.write("")
        riepilogo = f"Fatto: {n_create} contatti creati, {n_update} aggiornati, {n_err} errori."
        if dry:
            riepilogo = "[DRY-RUN] " + riepilogo + " (nessun salvataggio)"
        if n_err:
            self.stdout.write(self.style.WARNING(riepilogo))
        else:
            self.stdout.write(self.style.SUCCESS(riepilogo))
