"""Genera al volo i template Excel per gli import."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment


def build_template_servizi_workbook():
    """Crea e ritorna un openpyxl Workbook col template Excel per importa_servizi."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Servizi"

    headers = [
        ("evento_slug", "OBBLIGATORIO. Slug dell'evento, es. AITEB2026"),
        ("code", "OBBLIGATORIO. Codice univoco per evento, es. TAVOLO_STD"),
        ("nome_it", "OBBLIGATORIO. Nome in italiano"),
        ("nome_en", "Opzionale. Nome in inglese"),
        ("descrizione_it", "Opzionale. Descrizione italiano"),
        ("descrizione_en", "Opzionale. Descrizione inglese"),
        ("categoria", "Opzionale. Categoria libera (es. arredo, grafica, tecnico)"),
        ("categoria_contabile", "Default 'altro'. Valori: viaggio_partecipanti, "
         "viaggio_relatori, affitto_sala, stand, coffee_break, scheda_tecnica, "
         "quota_iscrizione, altro"),
        ("prezzo_base", "OBBLIGATORIO. Prezzo unitario in euro (es. 100.00)"),
        ("iva_percento", "IVA in percentuale (default 22)"),
        ("attivo", "s/n (default s)"),
        ("quantita_max", "Quantita' massima vendibile per contratto (vuoto = illimitata)"),
        ("quantita_totale", "Numero massimo disponibile in totale (pezzi esistenti, es. 1 "
         "per uno stand unico). Vuoto = illimitato."),
        ("ordine", "Ordine di visualizzazione (numero, default 0)"),
        ("pricing_mode", "Default 'fixed'. Valori: fixed, quantity, tiered"),
        ("genera_scadenze", "s/n (default n). Se 's', vendere il servizio crea le scadenze "
         "materiali (servono i modelli di scadenza sul servizio)."),
        ("servizi_inclusi", "Opzionale. Codici di altri servizi dello STESSO evento separati "
         "da virgola: vengono aggiunti a 0 EUR quando questo servizio viene venduto."),
        ("immagine", "Opzionale. PERCORSO COMPLETO del file foto, anche in stile Windows "
         "(es. C:\\Users\\morin\\foto\\stand.jpg) - viene convertito da solo. In alternativa "
         "solo il nome file (es. stand.jpg) importando con: --immagini <cartella>."),
    ]

    header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for i, (name, comment) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.comment = Comment(comment, "Sponsor Manager")

    esempio1 = ["AITEB2026", "TAVOLO_DEMO", "Tavolo standard", "Standard table",
                "Tavolo 80x80 cm, bianco", "80x80 cm white table",
                "arredo", "altro", 100.00, 22, "s", "", 10, "fixed", "n", "FARETTO_DEMO"]
    esempio2 = ["AITEB2026", "FARETTO_DEMO", "Faretto LED", "LED spotlight",
                "Faretto LED 30W", "LED spotlight 30W",
                "tecnico", "altro", 25.00, 22, "s", 50, 20, "fixed", "s", ""]
    for col, v in enumerate(esempio1, start=1):
        ws.cell(row=2, column=col, value=v)
    for col, v in enumerate(esempio2, start=1):
        ws.cell(row=3, column=col, value=v)

    larghezze = [14, 18, 22, 22, 28, 28, 14, 22, 14, 14, 8, 14, 10, 14, 16, 30]
    for i, w in enumerate(larghezze, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws2 = wb.create_sheet("Istruzioni")
    istr = [
        "IMPORT SERVIZI - Istruzioni",
        "",
        "1. Le PRIME 2 RIGHE del foglio 'Servizi' sono esempi: cancellali o sovrascrivili.",
        "2. Compila una riga per ogni servizio da importare/aggiornare.",
        "3. Colonne OBBLIGATORIE: evento_slug, code, nome_it, prezzo_base.",
        "4. evento_slug: lo slug dell'evento gia' esistente (es. AITEB2026).",
        "5. code: univoco per evento. Se esiste -> AGGIORNA, sennò -> CREA.",
        "6. categoria_contabile valida (default 'altro'):",
        "   viaggio_partecipanti / viaggio_relatori / affitto_sala / stand /",
        "   coffee_break / scheda_tecnica / quota_iscrizione / altro",
        "7. pricing_mode valida (default 'fixed'): fixed / quantity / tiered",
        "8. attivo: s/n (default s). Numeri: usa il punto o la virgola decimale.",
        "9. genera_scadenze: s/n. Se 's', vendere il servizio crea le scadenze materiali.",
        "10. servizi_inclusi: codici di altri servizi dello stesso evento (separati da virgola).",
        "    Vengono aggiunti a 0 EUR quando questo servizio entra in un contratto.",
        "11. immagine: due modi.",
        "    a) PERCORSO COMPLETO nella cella (consigliato, funziona anche dall'import web):",
        "       es. C:\\Users\\morin\\foto\\stand.jpg  (i percorsi Windows si convertono da soli).",
        "    b) solo il NOME del file (es. stand.jpg) + cartella da riga di comando:",
        "       python manage.py importa_servizi --file <file.xlsx> --immagini <cartella>",
        "",
        "Lancio:",
        "  python manage.py importa_servizi --file <percorso_file.xlsx> --dry-run",
        "  (verifica cosa farebbe, poi togli --dry-run per eseguire davvero)",
    ]
    for r, riga in enumerate(istr, start=1):
        ws2.cell(row=r, column=1, value=riga)
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws2.column_dimensions["A"].width = 80
    return wb


def build_template_catalogo_workbook():
    """Template Excel per importa_catalogo (CATALOGO GENERALE, senza evento)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.comments import Comment

    wb = Workbook()
    ws = wb.active
    ws.title = "Catalogo"

    headers = [
        ("code", "OBBLIGATORIO. Codice univoco nel catalogo, es. TAVOLO_STD"),
        ("nome_it", "OBBLIGATORIO. Nome in italiano"),
        ("nome_en", "Opzionale. Nome in inglese"),
        ("descrizione_it", "Opzionale. Descrizione italiano"),
        ("descrizione_en", "Opzionale. Descrizione inglese"),
        ("categoria", "Opzionale. Nome categoria (creata se non esiste), es. Arredo"),
        ("categoria_contabile", "Default 'altro'. Valori: viaggio_partecipanti, "
         "viaggio_relatori, affitto_sala, stand, coffee_break, scheda_tecnica, "
         "quota_iscrizione, altro"),
        ("prezzo_base", "OBBLIGATORIO. Prezzo base di default in euro (es. 100.00)"),
        ("iva_percento", "IVA in percentuale (default 22)"),
        ("attivo", "s/n (default s)"),
        ("quantita_max", "Quantita' massima per contratto (vuoto = illimitata)"),
        ("ordine", "Ordine di visualizzazione (numero, default 0)"),
        ("pricing_mode", "Default 'fixed'. Valori: fixed, quantity, tiered"),
        ("genera_scadenze", "s/n (default n). Se 's', vendere il servizio crea le scadenze materiali."),
        ("self_service", "s/n (default n). Acquistabile dal cliente in self-service."),
        ("cutoff_giorni", "Giorni prima dell'evento entro cui si puo' acquistare self-service (vuoto = nessun limite)."),
        ("immagine", "Opzionale. PERCORSO COMPLETO del file foto, anche stile Windows "
         "(es. C:\\Users\\morin\\foto\\stand.jpg) - convertito da solo. Oppure solo il nome "
         "file importando da CLI con: --immagini <cartella>."),
    ]

    header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for i, (name, comment) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.comment = Comment(comment, "Sponsor Manager")

    esempio = ["TAVOLO_STD", "Tavolo standard", "Standard table",
               "Tavolo 80x80 cm, bianco", "80x80 cm white table",
               "Arredo", "altro", 100.00, 22, "s", "", 10, "fixed", "n", "n", "", ""]
    for col, v in enumerate(esempio, start=1):
        ws.cell(row=2, column=col, value=v)

    larghezze = [18, 22, 22, 28, 28, 16, 22, 14, 14, 8, 14, 10, 14, 16, 12, 14, 40]
    for i, w in enumerate(larghezze, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws2 = wb.create_sheet("Istruzioni")
    istr = [
        "IMPORT CATALOGO GENERALE - Istruzioni",
        "",
        "Questo importa nel CATALOGO MADRE (valido per tutti gli eventi), NON nel singolo evento.",
        "Dal catalogo poi crei i servizi di un evento scegliendo 'dal catalogo'.",
        "",
        "1. La RIGA 2 e' un esempio: cancellala o sovrascrivila.",
        "2. Colonne OBBLIGATORIE: code, nome_it, prezzo_base. (NESSUN evento_slug qui!)",
        "3. code: univoco nel catalogo. Se esiste -> AGGIORNA, sennò -> CREA.",
        "4. categoria_contabile (default 'altro'): viaggio_partecipanti / viaggio_relatori /",
        "   affitto_sala / stand / coffee_break / scheda_tecnica / quota_iscrizione / altro",
        "5. pricing_mode (default 'fixed'): fixed / quantity / tiered",
        "6. immagine: PERCORSO COMPLETO nella cella (anche C:\\... si converte da solo),",
        "   oppure solo nome file + da CLI: python manage.py importa_catalogo --file <f.xlsx> --immagini <cartella>",
    ]
    for r, riga in enumerate(istr, start=1):
        ws2.cell(row=r, column=1, value=riga)
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws2.column_dimensions["A"].width = 90
    return wb


def build_template_stand_workbook():
    """Crea e ritorna un openpyxl Workbook col template Excel per importa_stand."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.comments import Comment

    wb = Workbook()
    ws = wb.active
    ws.title = "Stand"

    headers = [
        ("evento_slug", "OBBLIGATORIO. Slug dell'evento, es. AITEB2026"),
        ("code", "OBBLIGATORIO. Codice univoco per evento, es. A-12"),
        ("blocco_code", "Opzionale. Codice del blocco (StandBlock) cui appartiene lo stand"),
        ("prezzo_base", "OBBLIGATORIO. Prezzo base in euro (es. 1500.00)"),
        ("larghezza_m", "Opzionale. Larghezza in metri (es. 3)"),
        ("profondita_m", "Opzionale. Profondita' in metri (es. 2)"),
        ("tipologia", "Opzionale. Testo libero (es. corner_premium, linear, island, custom)"),
        ("stato", "Default 'available'. Valori: available, reserved, assigned, unavailable"),
        ("allaccio_elettrico", "s/n (default n)"),
        ("potenza_kw", "Opzionale. Potenza in kW (es. 3)"),
        ("allaccio_idrico", "s/n (default n)"),
        ("internet", "s/n (default n)"),
        ("altezza_max_m", "Opzionale. Altezza massima in metri"),
        ("descrizione_preventivo", "Opzionale. Descrizione mostrata nel preventivo"),
    ]

    header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for i, (name, comment) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.comment = Comment(comment, "Sponsor Manager")

    esempio1 = ["AITEB2026", "A-01", "", 1500.00, 3, 2, "linear", "available",
                "s", 3, "n", "s", 2.5, "Stand lineare 3x2 m, fronte corridoio"]
    esempio2 = ["AITEB2026", "A-02", "", 2500.00, 4, 4, "island", "available",
                "s", 6, "s", "s", 3.0, "Isola 4x4 m, doppio fronte"]
    for col, v in enumerate(esempio1, start=1):
        ws.cell(row=2, column=col, value=v)
    for col, v in enumerate(esempio2, start=1):
        ws.cell(row=3, column=col, value=v)

    larghezze = [14, 10, 14, 14, 12, 12, 16, 14, 18, 12, 16, 10, 14, 32]
    for i, w in enumerate(larghezze, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws2 = wb.create_sheet("Istruzioni")
    istr = [
        "IMPORT STAND - Istruzioni",
        "",
        "1. Le PRIME 2 RIGHE del foglio 'Stand' sono esempi: cancellali o sovrascrivili.",
        "2. Compila una riga per ogni stand da importare/aggiornare.",
        "3. Colonne OBBLIGATORIE: evento_slug, code, prezzo_base.",
        "4. evento_slug: lo slug dell'evento gia' esistente (es. AITEB2026).",
        "5. code: univoco per evento. Se esiste -> AGGIORNA, sennò -> CREA.",
        "6. blocco_code: opzionale. Se valorizzato, lo stand viene collegato al blocco",
        "   (StandBlock) con quel codice NELLO STESSO evento. Se il blocco non esiste -> errore.",
        "7. stato valido (default 'available'): available / reserved / assigned / unavailable",
        "8. allaccio_elettrico / allaccio_idrico / internet: s/n (default n).",
        "9. Numeri: usa il punto o la virgola decimale.",
        "",
        "Lancio:",
        "  python manage.py importa_stand --file <percorso_file.xlsx> --dry-run",
        "  (verifica cosa farebbe, poi togli --dry-run per eseguire davvero)",
    ]
    for r, riga in enumerate(istr, start=1):
        ws2.cell(row=r, column=1, value=riga)
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws2.column_dimensions["A"].width = 80
    return wb


def _norm_bool_out(v):
    return 's' if v else 'n'


def export_servizi_workbook(event):
    """Workbook con i servizi esistenti di un evento (stesse colonne del template)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from catalog.models import Service

    wb = Workbook()
    ws = wb.active
    ws.title = "servizi"
    headers = ["evento_slug", "code", "nome_it", "nome_en", "descrizione_it",
               "descrizione_en", "categoria", "categoria_contabile", "prezzo_base",
               "iva_percento", "attivo", "quantita_max", "ordine", "pricing_mode",
               "genera_scadenze", "servizi_inclusi"]
    header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for i, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = header_fill
        c.font = header_font

    r = 2
    for s in Service.objects.filter(event=event).order_by('display_order', 'code'):
        nome = s.name if isinstance(s.name, dict) else {}
        desc = s.description if isinstance(s.description, dict) else {}
        ws.cell(row=r, column=1, value=event.slug)
        ws.cell(row=r, column=2, value=s.code)
        ws.cell(row=r, column=3, value=nome.get('it', ''))
        ws.cell(row=r, column=4, value=nome.get('en', ''))
        ws.cell(row=r, column=5, value=desc.get('it', ''))
        ws.cell(row=r, column=6, value=desc.get('en', ''))
        ws.cell(row=r, column=7, value=s.category or '')
        ws.cell(row=r, column=8, value=s.accounting_category or '')
        ws.cell(row=r, column=9, value=float(s.base_price) if s.base_price is not None else '')
        ws.cell(row=r, column=10, value=float(s.vat_rate) if s.vat_rate is not None else '')
        ws.cell(row=r, column=11, value=_norm_bool_out(s.is_active))
        ws.cell(row=r, column=12, value=s.max_quantity if s.max_quantity is not None else '')
        ws.cell(row=r, column=13, value=s.display_order or 0)
        ws.cell(row=r, column=14, value=s.pricing_mode or 'fixed')
        ws.cell(row=r, column=15, value=_norm_bool_out(s.triggers_deadlines))
        try:
            incl = ', '.join(sub.code for sub in s.included_services.all())
        except Exception:
            incl = ''
        ws.cell(row=r, column=16, value=incl)
        r += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
    return wb


def export_stand_workbook(event):
    """Workbook con gli stand esistenti di un evento (stesse colonne del template)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from venues.models import Stand

    wb = Workbook()
    ws = wb.active
    ws.title = "stand"
    headers = ["evento_slug", "code", "blocco_code", "prezzo_base", "larghezza_m",
               "profondita_m", "tipologia", "stato", "allaccio_elettrico", "potenza_kw",
               "allaccio_idrico", "internet", "altezza_max_m", "descrizione_preventivo"]
    header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for i, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = header_fill
        c.font = header_font

    r = 2
    for st in Stand.objects.filter(event=event).select_related('stand_block').order_by('code'):
        ws.cell(row=r, column=1, value=event.slug)
        ws.cell(row=r, column=2, value=st.code)
        ws.cell(row=r, column=3, value=st.stand_block.code if st.stand_block_id else '')
        ws.cell(row=r, column=4, value=float(st.base_price) if st.base_price is not None else '')
        ws.cell(row=r, column=5, value=float(st.width_meters) if st.width_meters is not None else '')
        ws.cell(row=r, column=6, value=float(st.depth_meters) if st.depth_meters is not None else '')
        ws.cell(row=r, column=7, value=st.stand_type or '')
        ws.cell(row=r, column=8, value=st.status or 'available')
        ws.cell(row=r, column=9, value=_norm_bool_out(st.has_power))
        ws.cell(row=r, column=10, value=float(st.power_kw) if st.power_kw is not None else '')
        ws.cell(row=r, column=11, value=_norm_bool_out(st.has_water))
        ws.cell(row=r, column=12, value=_norm_bool_out(st.has_internet))
        ws.cell(row=r, column=13, value=float(st.max_height_meters) if st.max_height_meters is not None else '')
        _qd = st.quote_description if isinstance(st.quote_description, dict) else {}
        ws.cell(row=r, column=14, value=_qd.get('it', '') or '')
        r += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
    return wb

def export_sponsor_workbook(sponsors):
    """Workbook con gli sponsor/clienti esistenti (stesse colonne del template
    import, così il file è ri-importabile). Il referente è il contatto principale
    (o il primo disponibile)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Sponsor"
    headers = ["ragione_sociale", "nome_commerciale", "partita_iva", "codice_fiscale",
               "codice_sdi", "pec", "indirizzo", "citta", "cap", "provincia", "paese",
               "settore", "sito_web", "note",
               "referente_nome", "referente_email", "referente_telefono", "referente_ruolo"]
    header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for i, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = header_fill
        c.font = header_font

    r = 2
    for s in sponsors:
        try:
            ref = s.contacts.filter(is_primary=True).first() or s.contacts.first()
        except Exception:
            ref = None
        vals = [
            s.legal_name or '', getattr(s, 'display_name', '') or '', s.vat_number or '',
            s.tax_code or '', s.sdi_code or '', s.pec_email or '', s.address_street or '',
            s.address_city or '', s.address_zip or '', s.address_province or '',
            s.address_country or '', getattr(s, 'industry', '') or '', s.website or '',
            s.notes or '',
            (ref.full_name if ref else ''), (ref.email if ref else ''),
            (getattr(ref, 'phone', '') if ref else ''), (getattr(ref, 'job_title', '') if ref else ''),
        ]
        for col, v in enumerate(vals, start=1):
            ws.cell(row=r, column=col, value=v)
        r += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
    return wb


def export_contatti_workbook(contatti):
    """Workbook con i contatti esistenti (stesse colonne del template import)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font

    ROLE_IT = {'signer': 'firmatario', 'marketing': 'marketing', 'finance': 'amministrazione',
               'operational': 'operativo', 'cc': 'cc', 'educational': 'educational'}
    wb = Workbook()
    ws = wb.active
    ws.title = "Contatti"
    headers = ["sponsor_partita_iva", "sponsor_ragione_sociale", "cognome", "nome", "email",
               "telefono", "ruolo_aziendale", "ruoli_funzionali", "principale",
               "consenso_marketing", "lingua", "note"]
    header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for i, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = header_fill
        c.font = header_font

    r = 2
    for ct in contatti:
        sp = ct.sponsor
        cognome = getattr(ct, 'last_name', '') or ''
        nome = getattr(ct, 'first_name', '') or ''
        if not (cognome or nome) and getattr(ct, 'full_name', ''):
            parts = ct.full_name.strip().split()
            if parts:
                cognome = parts[-1]
                nome = ' '.join(parts[:-1])
        ruoli = ', '.join(ROLE_IT.get(x, x) for x in (ct.roles or []))
        vals = [
            (sp.vat_number if sp else '') or '', (sp.legal_name if sp else '') or '',
            cognome, nome, ct.email or '', getattr(ct, 'phone', '') or '',
            getattr(ct, 'job_title', '') or '', ruoli,
            's' if getattr(ct, 'is_primary', False) else 'n',
            's' if getattr(ct, 'marketing_consent', False) else 'n',
            getattr(ct, 'preferred_language', 'it') or 'it', getattr(ct, 'notes', '') or '',
        ]
        for col, v in enumerate(vals, start=1):
            ws.cell(row=r, column=col, value=v)
        r += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
    return wb


def build_template_sponsor_workbook():
    """Crea e ritorna un Workbook col template Excel per importa_sponsor."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sponsor"

    headers = [
        ("ragione_sociale", "OBBLIGATORIO. Ragione sociale / nome azienda."),
        ("nome_commerciale", "Opzionale. Nome commerciale se diverso."),
        ("partita_iva", "Consigliato. Se presente e gia' esistente, lo sponsor viene AGGIORNATO."),
        ("codice_fiscale", "Opzionale."),
        ("codice_sdi", "Opzionale. Codice destinatario SDI (max 7)."),
        ("pec", "Opzionale. Indirizzo PEC."),
        ("indirizzo", "Opzionale. Via e civico sede legale."),
        ("citta", "Opzionale."),
        ("cap", "Opzionale."),
        ("provincia", "Opzionale. Sigla provincia (es. BO)."),
        ("paese", "Default IT. Sigla nazione, 2 lettere."),
        ("settore", "Opzionale. Settore merceologico."),
        ("sito_web", "Opzionale. URL completo (https://...)."),
        ("note", "Opzionale."),
        ("referente_nome", "Opzionale. Nome e cognome del referente (crea un contatto)."),
        ("referente_email", "Opzionale. Email del referente (serve INSIEME al nome)."),
        ("referente_telefono", "Opzionale."),
        ("referente_ruolo", "Opzionale. Es. Responsabile marketing."),
    ]
    header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for i, (name, comment) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.comment = Comment(comment, "Sponsor Manager")

    esempio1 = ["Rossi Pharma S.p.A.", "Rossi Pharma", "01234567890", "", "", "rossipharma@pec.it",
                "Via Roma 1", "Bologna", "40100", "BO", "IT", "Farmaceutico",
                "https://www.rossipharma.it", "",
                "Maria Bianchi", "maria.bianchi@rossipharma.it", "051 123456", "Responsabile marketing"]
    esempio2 = ["Verdi Medical S.r.l.", "", "09876543210", "", "", "",
                "Viale Italia 22", "Milano", "20100", "MI", "IT", "Dispositivi medici", "", "Cliente storico",
                "", "", "", ""]
    for col, v in enumerate(esempio1, start=1):
        ws.cell(row=2, column=col, value=v)
    for col, v in enumerate(esempio2, start=1):
        ws.cell(row=3, column=col, value=v)

    larghezze = [26, 20, 16, 16, 12, 24, 24, 16, 8, 10, 8, 18, 26, 24, 22, 28, 16, 22]
    for i, w in enumerate(larghezze, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws2 = wb.create_sheet("Istruzioni")
    istr = [
        "IMPORT SPONSOR / CLIENTI - Istruzioni",
        "",
        "1. Le PRIME 2 RIGHE del foglio 'Sponsor' sono esempi: cancellali o sovrascrivili.",
        "2. Compila una riga per ogni sponsor/cliente da importare o aggiornare.",
        "3. Unica colonna OBBLIGATORIA: ragione_sociale.",
        "4. RICONOSCIMENTO: se metti la partita_iva e ne esiste gia' una uguale lo sponsor",
        "   viene AGGIORNATO; altrimenti si cerca per ragione sociale; sennò si CREA.",
        "5. In aggiornamento le celle VUOTE NON cancellano i dati gia' presenti.",
        "6. Referente: compila SIA referente_nome SIA referente_email per creare il contatto.",
        "   Diventa 'principale' se lo sponsor non ne ha gia' uno.",
        "   NB: l'import NON crea l'accesso al portale: lo abiliti tu dall'admin quando vuoi.",
        "",
        "Consiglio: lascia spuntato 'Solo anteprima' al primo caricamento per vedere cosa farebbe.",
    ]
    for r, riga in enumerate(istr, start=1):
        ws2.cell(row=r, column=1, value=riga)
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws2.column_dimensions["A"].width = 80
    return wb

def build_template_contatti_workbook():
    """Crea e ritorna un Workbook col template Excel per importa_contatti."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Contatti"

    headers = [
        ("sponsor_partita_iva", "Consigliato. P.IVA dello sponsor a cui collegare il contatto."),
        ("sponsor_ragione_sociale", "Alternativa: ragione sociale dello sponsor (se non metti la P.IVA)."),
        ("cognome", "OBBLIGATORIO. Cognome del contatto."),
        ("nome", "Nome del contatto."),
        ("email", "OBBLIGATORIO. Email del contatto."),
        ("telefono", "Opzionale."),
        ("ruolo_aziendale", "Opzionale. Es. Responsabile marketing."),
        ("ruoli_funzionali", "Opzionale. Uno o piu' tra: firmatario, marketing, amministrazione, "
         "operativo, cc, educational (separati da virgola)."),
        ("principale", "s/n. Se 's' diventa il contatto principale dello sponsor."),
        ("consenso_marketing", "s/n."),
        ("lingua", "it/en (default it)."),
        ("note", "Opzionale."),
    ]
    header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for i, (name, comment) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.comment = Comment(comment, "Sponsor Manager")

    esempio1 = ["01234567890", "Rossi Pharma S.p.A.", "Bianchi", "Maria", "maria.bianchi@rossipharma.it",
                "051 123456", "Responsabile marketing", "marketing, cc", "s", "s", "it", ""]
    esempio2 = ["", "Verdi Medical S.r.l.", "Verdi", "Luca", "luca.verdi@verdimedical.it",
                "", "Amministrazione", "amministrazione, firmatario", "n", "n", "it", "Referente fatture"]
    for col, v in enumerate(esempio1, start=1):
        ws.cell(row=2, column=col, value=v)
    for col, v in enumerate(esempio2, start=1):
        ws.cell(row=3, column=col, value=v)

    larghezze = [18, 26, 18, 18, 30, 16, 24, 34, 12, 16, 10, 24]
    for i, w in enumerate(larghezze, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws2 = wb.create_sheet("Istruzioni")
    istr = [
        "IMPORT CONTATTI - Istruzioni",
        "",
        "1. Le PRIME 2 RIGHE del foglio 'Contatti' sono esempi: cancellali o sovrascrivili.",
        "2. Compila una riga per ogni contatto da importare o aggiornare.",
        "3. Colonne OBBLIGATORIE: cognome, email (il nome e' consigliato).",
        "4. COLLEGAMENTO ALLO SPONSOR: indica sponsor_partita_iva (consigliato) oppure",
        "   sponsor_ragione_sociale. Lo sponsor deve gia' esistere (importalo prima se serve).",
        "5. RICONOSCIMENTO: il contatto e' identificato da sponsor + email. Se esiste -> AGGIORNA,",
        "   altrimenti -> CREA. In aggiornamento le celle vuote NON cancellano i dati presenti.",
        "6. ruoli_funzionali: uno o piu' valori separati da virgola tra:",
        "   firmatario, marketing, amministrazione, operativo, cc, educational.",
        "7. principale: 's' rende il contatto il principale dello sponsor (gli altri vengono declassati).",
        "8. lingua: it/en (default it). consenso_marketing: s/n.",
        "   NB: l'import NON crea l'accesso al portale: lo abiliti tu dall'admin quando vuoi.",
        "",
        "Consiglio: lascia spuntato 'Solo anteprima' al primo caricamento per vedere cosa farebbe.",
    ]
    for r, riga in enumerate(istr, start=1):
        ws2.cell(row=r, column=1, value=riga)
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws2.column_dimensions["A"].width = 82
    return wb
