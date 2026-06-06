"""Importa stand da un file Excel."""
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


COLONNE_OBBLIGATORIE = ["evento_slug", "code", "prezzo_base"]
STATI_VALIDI = {"available", "reserved", "assigned", "unavailable"}


def _norm_bool(v, default=False):
    if v is None or v == "":
        return default
    s = str(v).strip().lower()
    if s in {"s", "si", "sì", "yes", "y", "1", "true", "x", "✓"}:
        return True
    if s in {"n", "no", "0", "false", "-"}:
        return False
    return default


def _norm_dec(v, campo):
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v).replace(",", "."))
    except (InvalidOperation, ValueError):
        raise ValueError(f"{campo}: valore non numerico {v!r}")


class Command(BaseCommand):
    help = "Importa o aggiorna Stand da un file Excel (vedi template_stand.xlsx)."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Percorso del file Excel.")
        parser.add_argument("--dry-run", action="store_true",
                            help="Simula senza scrivere nulla nel DB.")

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError("openpyxl non installato. pip install openpyxl")
        from venues.models import Stand, StandBlock
        from events.models import Event

        path = Path(opts["file"]).expanduser()
        if not path.exists():
            raise CommandError(f"File non trovato: {path}")

        dry = opts["dry_run"]
        self.stdout.write(f"Lettura {path.name}" + (" [DRY-RUN]" if dry else ""))

        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f"Impossibile aprire il file Excel: {e}")

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("Il foglio e' vuoto.")

        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        mancanti = [c for c in COLONNE_OBBLIGATORIE if c not in header]
        if mancanti:
            raise CommandError(f"Colonne obbligatorie mancanti: {mancanti}")
        col_idx = {h: i for i, h in enumerate(header)}

        eventi_cache = {}
        blocchi_cache = {}

        def trova_evento(slug):
            if slug not in eventi_cache:
                eventi_cache[slug] = Event.objects.filter(slug=slug).first()
            return eventi_cache[slug]

        def trova_blocco(evento, code):
            key = (evento.id, code)
            if key not in blocchi_cache:
                blocchi_cache[key] = StandBlock.objects.filter(
                    event=evento, code=code).first()
            return blocchi_cache[key]

        n_create = n_update = n_err = n_dup = 0
        errori = []
        seen_codes = set()  # (evento_slug, code) già visti NEL FILE

        for n_riga, row in enumerate(rows[1:], start=2):
            if all(v is None or v == "" for v in row):
                continue

            def G(nome):
                i = col_idx.get(nome)
                return row[i] if i is not None and i < len(row) else None

            try:
                evento_slug = str(G("evento_slug") or "").strip()
                code = str(G("code") or "").strip()
                prezzo = _norm_dec(G("prezzo_base"), "prezzo_base")
                if not evento_slug or not code or prezzo is None:
                    raise ValueError("campi obbligatori mancanti (evento_slug, code, prezzo_base)")

                evento = trova_evento(evento_slug)
                if not evento:
                    raise ValueError(f"evento '{evento_slug}' non trovato")

                # codice duplicato NEL FILE: ogni stand deve avere un codice unico
                # per evento, altrimenti le righe si sovrascrivono a vicenda.
                key_dup = (evento_slug, code)
                if key_dup in seen_codes:
                    n_dup += 1
                    self.stdout.write(self.style.WARNING(
                        f"  {n_riga:>4}: ⚠ codice '{code}' DUPLICATO nel file su "
                        f"{evento_slug} - riga SALTATA (ogni stand serve un codice unico)"))
                    continue
                seen_codes.add(key_dup)

                # blocco opzionale
                blocco = None
                blocco_code = str(G("blocco_code") or "").strip()
                if blocco_code:
                    blocco = trova_blocco(evento, blocco_code)
                    if not blocco:
                        raise ValueError(
                            f"blocco '{blocco_code}' non trovato nell'evento '{evento_slug}'")

                stato = str(G("stato") or "available").strip().lower() or "available"
                if stato not in STATI_VALIDI:
                    raise ValueError(f"stato '{stato}' non valido. Validi: {sorted(STATI_VALIDI)}")

                larghezza = _norm_dec(G("larghezza_m"), "larghezza_m")
                profondita = _norm_dec(G("profondita_m"), "profondita_m")
                tipologia = str(G("tipologia") or "").strip()
                potenza = _norm_dec(G("potenza_kw"), "potenza_kw")
                altezza = _norm_dec(G("altezza_max_m"), "altezza_max_m")
                ha_elettrico = _norm_bool(G("allaccio_elettrico"))
                ha_idrico = _norm_bool(G("allaccio_idrico"))
                ha_internet = _norm_bool(G("internet"))
                descr = str(G("descrizione_preventivo") or "").strip()

                if dry:
                    esiste = Stand.objects.filter(event=evento, code=code).exists()
                    azione = "AGGIORNEREBBE" if esiste else "CREEREBBE"
                    self.stdout.write(f"  {n_riga:>4}: {azione} '{code}' su {evento_slug}"
                                      + (f" (blocco {blocco_code})" if blocco_code else ""))
                    if esiste:
                        n_update += 1
                    else:
                        n_create += 1
                    continue

                campi = {
                    "stand_block": blocco,
                    "base_price": prezzo,
                    "status": stato,
                    "stand_type": tipologia,
                    "has_power": ha_elettrico,
                    "has_water": ha_idrico,
                    "has_internet": ha_internet,
                    "quote_description": ({"it": descr} if descr else {}),
                }
                if larghezza is not None:
                    campi["width_meters"] = larghezza
                if profondita is not None:
                    campi["depth_meters"] = profondita
                if potenza is not None:
                    campi["power_kw"] = potenza
                if altezza is not None:
                    campi["max_height_meters"] = altezza

                with transaction.atomic():
                    stand, creato = Stand.objects.get_or_create(
                        event=evento, code=code, defaults=campi)
                    if creato:
                        n_create += 1
                        self.stdout.write(f"  {n_riga:>4}: + CREATO '{code}' su {evento_slug}")
                    else:
                        gia_assegnato = stand.contracts.exists()
                        if gia_assegnato:
                            # PROTEZIONE: non tocco base_price, status, stand_block
                            protetti = {'base_price', 'status', 'stand_block'}
                            for k, v in campi.items():
                                if k in protetti:
                                    continue
                                setattr(stand, k, v)
                            stand.save()
                            n_update += 1
                            self.stdout.write(self.style.WARNING(
                                f"  {n_riga:>4}: ~ AGGIORNATO (PROTETTO: ha contratti) '{code}' "
                                f"su {evento_slug} - prezzo/stato/blocco NON modificati"))
                        else:
                            for k, v in campi.items():
                                setattr(stand, k, v)
                            stand.save()
                            n_update += 1
                            self.stdout.write(f"  {n_riga:>4}: ~ AGGIORNATO '{code}' su {evento_slug}")

            except Exception as e:
                n_err += 1
                msg = f"riga {n_riga}: ERRORE {e}"
                errori.append(msg)
                self.stdout.write(self.style.ERROR("  " + msg))

        self.stdout.write("")
        riepilogo = f"Fatto: {n_create} creati, {n_update} aggiornati, {n_err} errori."
        if n_dup:
            riepilogo += f" {n_dup} righe saltate per codice duplicato."
        if dry:
            riepilogo = "[DRY-RUN] " + riepilogo + " (nessun salvataggio)"
        if n_err:
            self.stdout.write(self.style.WARNING(riepilogo))
        else:
            self.stdout.write(self.style.SUCCESS(riepilogo))
