"""Importa servizi nel CATALOGO (indipendente dall'evento) da un file Excel."""
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify


COLONNE_RICHIESTE = [
    "code", "nome_it", "nome_en", "descrizione_it", "descrizione_en",
    "categoria", "categoria_contabile", "prezzo_base", "iva_percento", "attivo",
    "quantita_max", "ordine", "pricing_mode", "genera_scadenze",
    "self_service", "cutoff_giorni", "immagine",
]
COLONNE_OBBLIGATORIE = ["code", "nome_it", "prezzo_base"]

ACCOUNTING_VALIDI = {
    "viaggio_partecipanti", "viaggio_relatori", "affitto_sala", "stand",
    "coffee_break", "scheda_tecnica", "quota_iscrizione", "altro",
}
PRICING_VALIDI = {"fixed", "quantity", "tiered"}


def _norm_bool(v, default=True):
    if v is None or v == "":
        return default
    s = str(v).strip().lower()
    if s in {"s", "si", "si'", "yes", "y", "1", "true", "x"}:
        return True
    if s in {"n", "no", "0", "false", "-"}:
        return False
    return default


def _norm_dec(v, campo):
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v).replace(",", "."))
    except (InvalidOperation, ValueError):
        raise ValueError(f"{campo}: valore non numerico {v!r}")


def _norm_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).replace(",", ".")))
    except ValueError:
        raise ValueError(f"valore non intero: {v!r}")


class Command(BaseCommand):
    help = "Importa/aggiorna il CATALOGO servizi da Excel (vedi template_catalogo.xlsx)."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Percorso del file Excel.")
        parser.add_argument("--immagini", default="",
                            help="Cartella con le immagini (colonna 'immagine').")
        parser.add_argument("--dry-run", action="store_true",
                            help="Simula senza scrivere nulla nel DB.")

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError("openpyxl non installato. Installa con: pip install openpyxl")
        from catalog.models import CatalogService, ServiceCategory

        path = Path(opts["file"]).expanduser()
        if not path.exists():
            raise CommandError(f"File non trovato: {path}")

        dry = opts["dry_run"]
        self.stdout.write(f"Lettura {path.name}" + (" [DRY-RUN]" if dry else ""))

        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f"Impossibile aprire il file Excel: {e}")

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("Il foglio e' vuoto.")

        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        mancanti = [c for c in COLONNE_OBBLIGATORIE if c not in header]
        if mancanti:
            raise CommandError(
                f"Colonne obbligatorie mancanti: {mancanti}. Attese: {COLONNE_RICHIESTE}"
            )
        col_idx = {h: i for i, h in enumerate(header)}

        cat_cache = {}

        def trova_o_crea_categoria(nome):
            nome = (nome or "").strip()
            if not nome:
                return None
            key = nome.lower()
            if key in cat_cache:
                return cat_cache[key]
            code = slugify(nome)[:50] or "categoria"
            obj, _ = ServiceCategory.objects.get_or_create(
                code=code, defaults={"name": nome, "is_active": True}
            )
            cat_cache[key] = obj
            return obj

        cartella_img = (opts.get("immagini") or "").strip()
        n_create = n_update = n_err = 0

        for n_riga, row in enumerate(rows[1:], start=2):
            if all(v is None or v == "" for v in row):
                continue

            def G(nome):
                i = col_idx.get(nome)
                return row[i] if i is not None and i < len(row) else None

            try:
                code = str(G("code") or "").strip()
                nome_it = str(G("nome_it") or "").strip()
                prezzo = _norm_dec(G("prezzo_base"), "prezzo_base")
                if not code or not nome_it or prezzo is None:
                    raise ValueError("campi obbligatori mancanti (code, nome_it, prezzo_base)")

                nome_en = str(G("nome_en") or "").strip()
                desc_it = str(G("descrizione_it") or "").strip()
                desc_en = str(G("descrizione_en") or "").strip()
                cat_nome = str(G("categoria") or "").strip()
                acc_cat = str(G("categoria_contabile") or "").strip() or "altro"
                if acc_cat not in ACCOUNTING_VALIDI:
                    raise ValueError(
                        f"categoria_contabile '{acc_cat}' non valida. Valide: {sorted(ACCOUNTING_VALIDI)}"
                    )
                iva = _norm_dec(G("iva_percento"), "iva_percento")
                attivo = _norm_bool(G("attivo"), default=True)
                qmax = _norm_int(G("quantita_max"))
                ordine = _norm_int(G("ordine")) or 0
                pmode = str(G("pricing_mode") or "fixed").strip().lower() or "fixed"
                if pmode not in PRICING_VALIDI:
                    raise ValueError(f"pricing_mode '{pmode}' non valido. Validi: {sorted(PRICING_VALIDI)}")
                genera = _norm_bool(G("genera_scadenze"), default=False)
                self_serv = _norm_bool(G("self_service"), default=False)
                cutoff = _norm_int(G("cutoff_giorni"))
                immagine_raw = str(G("immagine") or "").strip()

                name_json = {"it": nome_it}
                if nome_en:
                    name_json["en"] = nome_en
                desc_json = {}
                if desc_it:
                    desc_json["it"] = desc_it
                if desc_en:
                    desc_json["en"] = desc_en

                if dry:
                    esiste = CatalogService.objects.filter(code=code).exists()
                    azione = "AGGIORNEREBBE" if esiste else "CREEREBBE"
                    self.stdout.write(f"  {n_riga:>4}: {azione} '{code}'")
                    if esiste:
                        n_update += 1
                    else:
                        n_create += 1
                    continue

                categoria = trova_o_crea_categoria(cat_nome)

                with transaction.atomic():
                    svc, creato = CatalogService.objects.get_or_create(
                        code=code,
                        defaults={
                            "name": name_json,
                            "description": desc_json,
                            "category": categoria,
                            "accounting_category": acc_cat,
                            "pricing_mode": pmode,
                            "base_price": prezzo,
                            "vat_rate": iva if iva is not None else Decimal("22.00"),
                            "is_active": attivo,
                            "max_quantity": qmax,
                            "display_order": ordine,
                            "triggers_deadlines": genera,
                            "is_self_purchasable": self_serv,
                            "self_purchase_cutoff_days": cutoff,
                        },
                    )
                    if creato:
                        n_create += 1
                        self.stdout.write(f"  {n_riga:>4}: + CREATO '{code}'")
                    else:
                        svc.name = name_json
                        if desc_json:
                            svc.description = desc_json
                        svc.category = categoria
                        svc.accounting_category = acc_cat
                        svc.pricing_mode = pmode
                        svc.base_price = prezzo
                        if iva is not None:
                            svc.vat_rate = iva
                        svc.is_active = attivo
                        svc.max_quantity = qmax
                        svc.display_order = ordine
                        svc.triggers_deadlines = genera
                        svc.is_self_purchasable = self_serv
                        svc.self_purchase_cutoff_days = cutoff
                        svc.save()
                        n_update += 1
                        self.stdout.write(f"  {n_riga:>4}: ~ AGGIORNATO '{code}'")

                if immagine_raw and not dry:
                    _cand = Path(immagine_raw)
                    if _cand.is_file():
                        _img = _cand
                    elif cartella_img and (Path(cartella_img) / immagine_raw).is_file():
                        _img = Path(cartella_img) / immagine_raw
                    else:
                        _img = None
                    if _img is not None:
                        from django.core.files import File as _DjFile
                        with open(_img, "rb") as _fh:
                            svc.image.save(_img.name, _DjFile(_fh), save=True)
                        self.stdout.write(f"  {n_riga:>4}:   immagine agganciata: {_img.name}")
                    else:
                        self.stdout.write(self.style.WARNING(
                            f"  {n_riga:>4}:   immagine NON trovata: {immagine_raw}"))

            except Exception as e:
                n_err += 1
                self.stdout.write(self.style.ERROR(f"  riga {n_riga}: ERRORE {e}"))

        self.stdout.write("")
        riepilogo = f"Fatto: {n_create} create, {n_update} aggiornate, {n_err} errori."
        if dry:
            riepilogo = "[DRY-RUN] " + riepilogo + " (nessun salvataggio)"
        if n_err:
            self.stdout.write(self.style.WARNING(riepilogo))
        else:
            self.stdout.write(self.style.SUCCESS(riepilogo))
        if not dry and n_create:
            self.stdout.write(self.style.WARNING(
                "Catalogo aggiornato. Ricorda: per offrirli in un evento, "
                "spunta i servizi nella scheda dell'evento."
            ))
