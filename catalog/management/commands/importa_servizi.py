"""Importa servizi da un file Excel."""
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


COLONNE_RICHIESTE = [
    "evento_slug", "code", "nome_it", "nome_en", "descrizione_it", "descrizione_en",
    "categoria", "categoria_contabile", "prezzo_base", "iva_percento", "attivo",
    "quantita_max", "ordine", "pricing_mode",
]
COLONNE_OBBLIGATORIE = ["evento_slug", "code", "nome_it", "prezzo_base"]

ACCOUNTING_VALIDI = {
    "viaggio_partecipanti", "viaggio_relatori", "affitto_sala", "stand",
    "coffee_break", "scheda_tecnica", "quota_iscrizione", "altro",
}
PRICING_VALIDI = {"fixed", "quantity", "tiered"}


def _norm_bool(v, default=True):
    """s/si/sì/yes/y/1/True/x -> True; n/no/0/False/'' -> False."""
    if v is None or v == "":
        return default
    s = str(v).strip().lower()
    if s in {"s", "si", "sì", "yes", "y", "1", "true", "x", "✓"}:
        return True
    if s in {"n", "no", "0", "false", "-"}:
        return False
    return default


def _norm_dec(v, campo):
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v).replace(",", "."))
    except (InvalidOperation, ValueError):
        raise ValueError(f"{campo}: valore non numerico {v!r}")


def _norm_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).replace(",", ".")))
    except ValueError:
        raise ValueError(f"valore non intero: {v!r}")


class Command(BaseCommand):
    help = "Importa o aggiorna Service da un file Excel (vedi template_servizi.xlsx)."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Percorso del file Excel.")
        parser.add_argument("--dry-run", action="store_true",
                            help="Simula senza scrivere nulla nel DB.")

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError(
                "openpyxl non installato. Installa con: pip install openpyxl"
            )
        from catalog.models import Service
        from events.models import Event

        path = Path(opts["file"]).expanduser()
        if not path.exists():
            raise CommandError(f"File non trovato: {path}")

        dry = opts["dry_run"]
        self.stdout.write(f"Lettura {path.name}" + (" [DRY-RUN]" if dry else ""))

        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f"Impossibile aprire il file Excel: {e}")

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("Il foglio e' vuoto.")

        # 1) intestazione
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        mancanti = [c for c in COLONNE_OBBLIGATORIE if c not in header]
        if mancanti:
            raise CommandError(
                f"Colonne obbligatorie mancanti nell'intestazione: {mancanti}. "
                f"Atteso: {COLONNE_RICHIESTE}"
            )
        col_idx = {h: i for i, h in enumerate(header)}

        # 2) cache eventi per slug (1 query per slug)
        eventi_cache = {}

        def trova_evento(slug):
            if slug in eventi_cache:
                return eventi_cache[slug]
            ev = Event.objects.filter(slug=slug).first()
            eventi_cache[slug] = ev
            return ev

        # 3) elaborazione righe
        n_create = n_update = n_skip = n_err = 0
        errori = []

        for n_riga, row in enumerate(rows[1:], start=2):  # riga 2 e oltre (1 = header)
            if all(v is None or v == "" for v in row):
                continue  # riga totalmente vuota: salta

            def G(nome):
                i = col_idx.get(nome)
                return row[i] if i is not None and i < len(row) else None

            try:
                evento_slug = str(G("evento_slug") or "").strip()
                code = str(G("code") or "").strip()
                nome_it = str(G("nome_it") or "").strip()
                prezzo = _norm_dec(G("prezzo_base"), "prezzo_base")
                if not evento_slug or not code or not nome_it or prezzo is None:
                    raise ValueError(
                        "campi obbligatori mancanti (evento_slug, code, nome_it, prezzo_base)"
                    )

                evento = trova_evento(evento_slug)
                if not evento:
                    raise ValueError(f"evento '{evento_slug}' non trovato")

                nome_en = str(G("nome_en") or "").strip()
                desc_it = str(G("descrizione_it") or "").strip()
                desc_en = str(G("descrizione_en") or "").strip()
                cat = str(G("categoria") or "").strip()
                acc_cat = str(G("categoria_contabile") or "").strip() or "altro"
                if acc_cat not in ACCOUNTING_VALIDI:
                    raise ValueError(
                        f"categoria_contabile '{acc_cat}' non valida. "
                        f"Valide: {sorted(ACCOUNTING_VALIDI)}"
                    )
                iva = _norm_dec(G("iva_percento"), "iva_percento")
                attivo = _norm_bool(G("attivo"), default=True)
                qmax = _norm_int(G("quantita_max"))
                ordine = _norm_int(G("ordine")) or 0
                pmode = str(G("pricing_mode") or "fixed").strip().lower() or "fixed"
                if pmode not in PRICING_VALIDI:
                    raise ValueError(
                        f"pricing_mode '{pmode}' non valido. Validi: {sorted(PRICING_VALIDI)}"
                    )

                # campi name/description sono JSON multilingua
                name_json = {"it": nome_it}
                if nome_en:
                    name_json["en"] = nome_en
                desc_json = {}
                if desc_it:
                    desc_json["it"] = desc_it
                if desc_en:
                    desc_json["en"] = desc_en

                if dry:
                    esiste = Service.objects.filter(event=evento, code=code).exists()
                    azione = "AGGIORNEREBBE" if esiste else "CREEREBBE"
                    self.stdout.write(f"  {n_riga:>4}: {azione} '{code}' su {evento_slug}")
                    if esiste:
                        n_update += 1
                    else:
                        n_create += 1
                    continue

                with transaction.atomic():
                    svc, creato = Service.objects.get_or_create(
                        event=evento, code=code,
                        defaults={
                            "name": name_json,
                            "description": desc_json,
                            "category": cat,
                            "accounting_category": acc_cat,
                            "pricing_mode": pmode,
                            "base_price": prezzo,
                            "vat_rate": iva if iva is not None else Decimal("22.00"),
                            "is_active": attivo,
                            "max_quantity": qmax,
                            "display_order": ordine,
                        },
                    )
                    if creato:
                        n_create += 1
                        self.stdout.write(f"  {n_riga:>4}: + CREATO '{code}' su {evento_slug}")
                    else:
                        # AGGIORNA i campi
                        svc.name = name_json
                        if desc_json:
                            svc.description = desc_json
                        svc.category = cat
                        svc.accounting_category = acc_cat
                        svc.pricing_mode = pmode
                        svc.base_price = prezzo
                        if iva is not None:
                            svc.vat_rate = iva
                        svc.is_active = attivo
                        svc.max_quantity = qmax
                        svc.display_order = ordine
                        svc.save()
                        n_update += 1
                        self.stdout.write(f"  {n_riga:>4}: ~ AGGIORNATO '{code}' su {evento_slug}")

            except Exception as e:
                n_err += 1
                msg = f"riga {n_riga}: ERRORE {e}"
                errori.append(msg)
                self.stdout.write(self.style.ERROR("  " + msg))

        # 4) riepilogo
        self.stdout.write("")
        riepilogo = (
            f"Fatto: {n_create} create, {n_update} aggiornate, {n_err} errori."
        )
        if dry:
            riepilogo = "[DRY-RUN] " + riepilogo + " (nessun salvataggio)"
        if n_err:
            self.stdout.write(self.style.WARNING(riepilogo))
        else:
            self.stdout.write(self.style.SUCCESS(riepilogo))
        if not dry and n_create:
            self.stdout.write(self.style.WARNING(
                "Ricorda: i prezzi sono per-evento. Verifica i base_price nei nuovi servizi."
            ))
